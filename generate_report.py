"""
generate_report.py
==================
Generates a branded PDF Data Quality Report from current pipeline data.

Produces:
  - dq_report_YYYY-MM-DD.pdf   (main PDF report)
  - dq_report_YYYY-MM-DD.csv   (anomaly table)
  - dq_report_YYYY-MM-DD.xlsx  (multi-sheet Excel)

Usage:
    python generate_report.py
    python generate_report.py --output ./reports/
    python generate_report.py --pdf-only
    python generate_report.py --csv-only
    python generate_report.py --excel-only
"""

import json
import os
import sys
import csv
import sqlite3
from datetime import datetime, timezone

# ── Colours (brand palette) ───────────────────────────────────
BRAND_DARK   = (17/255,  24/255,  39/255)   # #111827
BRAND_BLUE   = (37/255,  99/255, 235/255)   # #2563EB
BRAND_RED    = (239/255, 68/255,  68/255)   # #EF4444
BRAND_AMBER  = (245/255,158/255,  11/255)   # #F59E0B
BRAND_GREEN  = (34/255, 197/255,  94/255)   # #22C55E
BRAND_GRAY   = (107/255,114/255, 128/255)   # #6B7280
BRAND_LIGHT  = (249/255,250/255, 251/255)   # #F9FAFB
WHITE        = (1, 1, 1)


def load_data():
    """Load all report data from JSON files and metadata.db."""
    base = "./web_dashboard/data"

    def load_json(fname):
        path = f"{base}/{fname}"
        if os.path.exists(path):
            with open(path) as f:
                return json.load(f)
        return {}

    summary    = load_json("summary.json")
    dq_scores  = load_json("dq_scores.json") or []
    anomalies  = load_json("anomalies.json") or []
    alert_log  = load_json("alert_log.json") or []
    charts     = load_json("charts.json")    or {}
    sources    = load_json("sources.json")   or {}

    return summary, dq_scores, anomalies, alert_log, charts, sources


# ════════════════════════════════════════════════════════════════
#  PDF REPORT
# ════════════════════════════════════════════════════════════════
def generate_pdf(output_path: str, summary, dq_scores, anomalies, alert_log, charts, sources):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        HRFlowable, PageBreak, KeepTogether
    )
    from reportlab.lib import colors
    from reportlab.graphics.shapes import Drawing, Rect, String, Line
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.charts.lineplots import LinePlot
    from reportlab.graphics import renderPDF

    W, H   = A4
    margin = 20 * mm

    doc = SimpleDocTemplate(
        output_path,
        pagesize=A4,
        leftMargin=margin, rightMargin=margin,
        topMargin=margin,  bottomMargin=margin,
        title="DQ Monitor — Data Quality Report",
        author="DQ Monitor",
    )

    # ── Colour helpers ────────────────────────────────────────
    def rgb(r,g,b): return colors.Color(r,g,b)

    C_DARK  = rgb(*BRAND_DARK)
    C_BLUE  = rgb(*BRAND_BLUE)
    C_RED   = rgb(*BRAND_RED)
    C_AMBER = rgb(*BRAND_AMBER)
    C_GREEN = rgb(*BRAND_GREEN)
    C_GRAY  = rgb(*BRAND_GRAY)
    C_LIGHT = rgb(*BRAND_LIGHT)

    def score_color(s):
        return C_RED if s < 70 else C_AMBER if s < 90 else C_GREEN

    def sev_color(s):
        return {
            "CRITICAL": C_RED,
            "HIGH":     C_AMBER,
            "MEDIUM":   C_AMBER,
            "LOW":      C_GREEN,
        }.get(s, C_GRAY)

    # ── Styles ────────────────────────────────────────────────
    styles = getSampleStyleSheet()

    H1 = ParagraphStyle("H1", fontSize=22, fontName="Helvetica-Bold",
                         textColor=C_DARK, spaceAfter=4)
    H2 = ParagraphStyle("H2", fontSize=14, fontName="Helvetica-Bold",
                         textColor=C_DARK, spaceBefore=14, spaceAfter=6)
    H3 = ParagraphStyle("H3", fontSize=11, fontName="Helvetica-Bold",
                         textColor=C_DARK, spaceBefore=8, spaceAfter=4)
    BODY = ParagraphStyle("BODY", fontSize=9, fontName="Helvetica",
                          textColor=C_GRAY, spaceAfter=4, leading=14)
    SMALL = ParagraphStyle("SMALL", fontSize=8, fontName="Helvetica",
                           textColor=C_GRAY, spaceAfter=2)
    MONO = ParagraphStyle("MONO", fontSize=8, fontName="Courier",
                          textColor=C_DARK, spaceAfter=2)
    EXPL = ParagraphStyle("EXPL", fontSize=8.5, fontName="Helvetica-Oblique",
                          textColor=C_GRAY, spaceAfter=3, leftIndent=8,
                          borderPad=4, leading=13)

    # ── Page callback for header/footer ───────────────────────
    run_date = summary.get("exported_at", datetime.now().isoformat())[:10]
    status   = summary.get("overall_status", "UNKNOWN")

    def on_page(canvas, doc):
        canvas.saveState()
        # Header bar
        canvas.setFillColorRGB(*BRAND_DARK)
        canvas.rect(0, H - 14*mm, W, 14*mm, fill=1, stroke=0)
        canvas.setFillColorRGB(1,1,1)
        canvas.setFont("Helvetica-Bold", 10)
        canvas.drawString(margin, H - 9*mm, "DQ Monitor")
        canvas.setFont("Helvetica", 8)
        canvas.setFillColorRGB(0.6, 0.7, 0.9)
        canvas.drawString(margin + 55, H - 9*mm, "Data Quality Report")
        canvas.setFillColorRGB(1,1,1)
        canvas.drawRightString(W - margin, H - 9*mm, run_date)
        # Footer
        canvas.setFillColorRGB(*BRAND_GRAY)
        canvas.setFont("Helvetica", 7)
        canvas.drawString(margin, 8*mm, f"Generated by DQ Monitor  ·  {run_date}")
        canvas.drawRightString(W - margin, 8*mm, f"Page {doc.page}")
        canvas.restoreState()

    # ── Build story ───────────────────────────────────────────
    story = []

    def section(title):
        story.append(Spacer(1, 4*mm))
        story.append(HRFlowable(width="100%", thickness=0.5,
                                color=C_LIGHT, spaceAfter=3))
        story.append(Paragraph(title, H2))

    # ── COVER ─────────────────────────────────────────────────
    story.append(Spacer(1, 12*mm))

    # Big status banner
    status_c = C_RED if status=="CRITICAL" else C_AMBER if status=="HIGH" else C_GREEN
    banner = Drawing(W - 2*margin, 28*mm)
    banner.add(Rect(0, 0, W - 2*margin, 28*mm,
                    fillColor=status_c, strokeColor=None, rx=4))
    banner.add(String((W - 2*margin)/2, 18*mm,
                       status,
                       textAnchor="middle",
                       fontName="Helvetica-Bold", fontSize=20,
                       fillColor=colors.white))
    banner.add(String((W - 2*margin)/2, 10*mm,
                       f"Overall System Status",
                       textAnchor="middle",
                       fontName="Helvetica", fontSize=9,
                       fillColor=colors.Color(1,1,1,0.8)))
    story.append(banner)
    story.append(Spacer(1, 6*mm))

    story.append(Paragraph("Data Quality Report", H1))
    story.append(Paragraph(
        f"Generated: {run_date}  ·  "
        f"Sources: {summary.get('source_count', 1)}  ·  "
        f"Tables monitored: {len(dq_scores)}",
        BODY
    ))

    # ── EXECUTIVE SUMMARY METRICS ──────────────────────────────
    story.append(Spacer(1, 6*mm))
    avg   = summary.get("avg_dq_score", 0)
    total = summary.get("total_anomalies", 0)
    crit  = summary.get("critical", 0)
    high  = summary.get("high", 0)
    snaps = summary.get("total_snapshots", 0)
    gx_p  = summary.get("gx_passed", 0)
    gx_f  = summary.get("gx_failed", 0)

    metric_data = [
        ["Avg DQ Score", "Anomalies", "Critical", "Checks Passed", "Snapshots"],
        [str(avg),       str(total),  str(crit),  str(gx_p),       str(snaps)],
    ]
    metric_tbl = Table(metric_data, colWidths=[(W - 2*margin)/5]*5)
    metric_tbl.setStyle(TableStyle([
        ("BACKGROUND",  (0,0), (-1,0), C_LIGHT),
        ("BACKGROUND",  (0,1), (-1,1), colors.white),
        ("FONTNAME",    (0,0), (-1,0), "Helvetica"),
        ("FONTNAME",    (0,1), (-1,1), "Helvetica-Bold"),
        ("FONTSIZE",    (0,0), (-1,0), 8),
        ("FONTSIZE",    (0,1), (-1,1), 18),
        ("TEXTCOLOR",   (0,0), (-1,0), C_GRAY),
        ("TEXTCOLOR",   (0,1), (0,1),  score_color(avg)),
        ("TEXTCOLOR",   (1,1), (1,1),  C_RED if total > 0 else C_GREEN),
        ("TEXTCOLOR",   (2,1), (2,1),  C_RED if crit > 0 else C_GREEN),
        ("TEXTCOLOR",   (3,1), (3,1),  C_GREEN),
        ("ALIGN",       (0,0), (-1,-1), "CENTER"),
        ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0,0), (-1,-1), [C_LIGHT, colors.white]),
        ("GRID",        (0,0), (-1,-1), 0.3, C_LIGHT),
        ("TOPPADDING",  (0,0), (-1,-1), 6),
        ("BOTTOMPADDING",(0,0),(-1,-1), 6),
        ("ROUNDEDCORNERS", [4]),
    ]))
    story.append(metric_tbl)

    # ── TABLE HEALTH SCORES ────────────────────────────────────
    section("Table Health Scores")

    if dq_scores:
        score_rows = [["Table", "Score", "Status", "Issues"]]
        for s in sorted(dq_scores, key=lambda x: x["score"]):
            score_rows.append([
                s["table"],
                str(s["score"]),
                s["status"],
                "; ".join(s.get("issues", [])[:2]) or "—",
            ])

        score_tbl = Table(score_rows,
                          colWidths=[90*mm, 25*mm, 30*mm, None])
        style = [
            ("BACKGROUND",  (0,0), (-1,0), C_DARK),
            ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
            ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",    (0,0), (-1,-1), 8),
            ("FONTNAME",    (0,1), (-1,-1), "Helvetica"),
            ("ALIGN",       (1,0), (2,-1),  "CENTER"),
            ("GRID",        (0,0), (-1,-1), 0.3, C_LIGHT),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, C_LIGHT]),
            ("TOPPADDING",  (0,0), (-1,-1), 5),
            ("BOTTOMPADDING",(0,0),(-1,-1), 5),
        ]
        for i, s in enumerate(dq_scores, start=1):
            c = score_color(s["score"])
            style.append(("TEXTCOLOR", (1,i), (1,i), c))
            style.append(("TEXTCOLOR", (2,i), (2,i), c))
            style.append(("FONTNAME",  (1,i), (2,i), "Helvetica-Bold"))

        score_tbl.setStyle(TableStyle(style))
        story.append(score_tbl)

    # ── ANOMALIES ──────────────────────────────────────────────
    section("Anomaly Findings")

    if not anomalies:
        story.append(Paragraph("✓ No anomalies detected in this run.", BODY))
    else:
        anom_rows = [["Table", "Severity", "Metric", "Today", "Expected", "Detector"]]
        for a in anomalies:
            anom_rows.append([
                a.get("table",""),
                a.get("severity",""),
                a.get("metric",""),
                str(a.get("today","—")),
                str(a.get("expected","—")),
                a.get("detector",""),
            ])

        anom_tbl = Table(anom_rows,
                         colWidths=[55*mm, 22*mm, 42*mm, 22*mm, 22*mm, 22*mm])
        astyle = [
            ("BACKGROUND",  (0,0), (-1,0), C_DARK),
            ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
            ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",    (0,0), (-1,-1), 7.5),
            ("FONTNAME",    (0,1), (-1,-1), "Helvetica"),
            ("ALIGN",       (1,0), (-1,-1), "CENTER"),
            ("GRID",        (0,0), (-1,-1), 0.3, C_LIGHT),
            ("ROWBACKGROUNDS", (0,1),(-1,-1),[colors.white, C_LIGHT]),
            ("TOPPADDING",  (0,0), (-1,-1), 4),
            ("BOTTOMPADDING",(0,0),(-1,-1), 4),
        ]
        for i, a in enumerate(anomalies, start=1):
            c = sev_color(a.get("severity",""))
            astyle.append(("TEXTCOLOR", (1,i), (1,i), c))
            astyle.append(("FONTNAME",  (1,i), (1,i), "Helvetica-Bold"))

        anom_tbl.setStyle(TableStyle(astyle))
        story.append(anom_tbl)

        # Explanations
        story.append(Spacer(1, 4*mm))
        story.append(Paragraph("Plain-English Explanations", H3))
        for a in anomalies:
            expl = a.get("explanation","")
            if expl:
                sev_c  = sev_color(a.get("severity",""))
                prefix = f"<b>{a.get('table','')} — {a.get('metric','')}:</b> "
                story.append(Paragraph(prefix + expl, EXPL))

    # ── ALERT LOG ──────────────────────────────────────────────
    if alert_log:
        section("Alert History")
        alert_rows = [["Sent at", "Channel", "Severity", "Summary"]]
        for a in alert_log[:20]:
            alert_rows.append([
                str(a.get("sent_at",""))[:19],
                a.get("channel",""),
                a.get("severity",""),
                a.get("summary","")[:60],
            ])
        alert_tbl = Table(alert_rows, colWidths=[40*mm, 22*mm, 22*mm, None])
        alert_tbl.setStyle(TableStyle([
            ("BACKGROUND",  (0,0), (-1,0), C_DARK),
            ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
            ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",    (0,0), (-1,-1), 7.5),
            ("FONTNAME",    (0,1), (-1,-1), "Helvetica"),
            ("GRID",        (0,0), (-1,-1), 0.3, C_LIGHT),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white, C_LIGHT]),
            ("TOPPADDING",  (0,0), (-1,-1), 4),
            ("BOTTOMPADDING",(0,0),(-1,-1), 4),
        ]))
        story.append(alert_tbl)

    # ── CONNECTED SOURCES ──────────────────────────────────────
    if sources:
        section("Connected Sources")
        for src_name, src in sources.items():
            story.append(Paragraph(
                f"<b>{src_name}</b>  ·  {src.get('dialect','').upper()}  ·  "
                f"{src.get('table_count',0)} table(s)  ·  "
                f"{src.get('description','')}",
                BODY
            ))

    # ── FOOTER NOTE ────────────────────────────────────────────
    story.append(Spacer(1, 8*mm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=C_LIGHT))
    story.append(Spacer(1, 3*mm))
    story.append(Paragraph(
        f"This report was automatically generated by DQ Monitor on {run_date}. "
        f"Data sourced from metadata.db and anomaly_report.json. "
        f"For questions, review the full dashboard at http://localhost:8080.",
        SMALL
    ))

    doc.build(story, onFirstPage=on_page, onLaterPages=on_page)
    print(f"  ✓ PDF:   {output_path}")


# ════════════════════════════════════════════════════════════════
#  CSV EXPORT
# ════════════════════════════════════════════════════════════════
def generate_csv(output_path: str, anomalies, dq_scores):
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["DQ Monitor — Anomaly Report"])
        w.writerow(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M")])
        w.writerow([])

        w.writerow(["Table Health Scores"])
        w.writerow(["Table", "Score", "Status", "Issues"])
        for s in dq_scores:
            w.writerow([s["table"], s["score"], s["status"],
                        "; ".join(s.get("issues",[]))])
        w.writerow([])

        w.writerow(["Anomalies"])
        w.writerow(["Detected At","Table","Severity","Metric",
                    "Detector","Today","Expected","Score","Explanation"])
        for a in anomalies:
            w.writerow([
                a.get("detected_at",""),
                a.get("table",""),
                a.get("severity",""),
                a.get("metric",""),
                a.get("detector",""),
                a.get("today",""),
                a.get("expected",""),
                a.get("score",""),
                a.get("explanation",""),
            ])

    print(f"  ✓ CSV:   {output_path}")


# ════════════════════════════════════════════════════════════════
#  EXCEL EXPORT
# ════════════════════════════════════════════════════════════════
def generate_excel(output_path: str, summary, dq_scores, anomalies, alert_log, charts):
    try:
        import openpyxl
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                      Border, Side, numbers)
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  ⚠ Excel skipped — run: pip install openpyxl")
        return

    wb = openpyxl.Workbook()

    # Colour fills
    def fill(hex_str): return PatternFill("solid", fgColor=hex_str)
    def font(bold=False, color="000000", size=10):
        return Font(bold=bold, color=color, size=size)

    DARK_FILL  = fill("111827")
    BLUE_FILL  = fill("2563EB")
    RED_FILL   = fill("FEF2F2")
    AMBER_FILL = fill("FFFBEB")
    GREEN_FILL = fill("F0FDF4")
    LIGHT_FILL = fill("F9FAFB")
    thin = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )

    def header_row(ws, row_num, headers, col_widths=None):
        for j, h in enumerate(headers, 1):
            c = ws.cell(row=row_num, column=j, value=h)
            c.fill = DARK_FILL
            c.font = font(bold=True, color="FFFFFF", size=9)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin
        if col_widths:
            for j, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(j)].width = w

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    # ── Sheet 1: Summary ──────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.row_dimensions[1].height = 30

    ws1["A1"] = "DQ Monitor — Data Quality Report"
    ws1["A1"].font = font(bold=True, size=14)
    ws1.merge_cells("A1:D1")

    ws1["A2"] = "Generated"
    ws1["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws1["A3"] = "Overall Status"
    ws1["B3"] = summary.get("overall_status","—")
    ws1["B3"].font = Font(bold=True, size=11,
                          color="B91C1C" if summary.get("overall_status")=="CRITICAL"
                          else "166534")
    ws1["A4"] = "Avg DQ Score"
    ws1["B4"] = summary.get("avg_dq_score", 0)
    ws1["A5"] = "Total Anomalies"
    ws1["B5"] = summary.get("total_anomalies", 0)
    ws1["A6"] = "Critical"
    ws1["B6"] = summary.get("critical", 0)
    ws1["A7"] = "Sources monitored"
    ws1["B7"] = summary.get("source_count", 1)
    ws1["A8"] = "Total snapshots"
    ws1["B8"] = summary.get("total_snapshots", 0)
    auto_width(ws1)

    # ── Sheet 2: DQ Scores ────────────────────────────────────
    ws2 = wb.create_sheet("DQ Scores")
    headers2 = ["Table", "Score", "Status", "Issues"]
    header_row(ws2, 1, headers2)
    for i, s in enumerate(dq_scores, 2):
        score_fill = (RED_FILL if s["score"] < 70
                      else AMBER_FILL if s["score"] < 90 else GREEN_FILL)
        cells = [s["table"], s["score"], s["status"],
                 "; ".join(s.get("issues",[]) or [""])]
        for j, val in enumerate(cells, 1):
            c = ws2.cell(row=i, column=j, value=val)
            c.border = thin
            c.alignment = Alignment(vertical="center")
            if j in (2, 3):
                c.fill = score_fill
                c.font = Font(bold=True, size=9)
        if i % 2 == 1:
            for j in range(1, 5):
                if ws2.cell(i,j).fill == PatternFill():
                    ws2.cell(i,j).fill = LIGHT_FILL
    auto_width(ws2)

    # ── Sheet 3: Anomalies ────────────────────────────────────
    ws3 = wb.create_sheet("Anomalies")
    headers3 = ["Detected At","Table","Severity","Metric",
                "Detector","Today","Expected","Score","Explanation"]
    header_row(ws3, 1, headers3)
    for i, a in enumerate(anomalies, 2):
        sev_fill = (RED_FILL   if a.get("severity")=="CRITICAL"
                    else AMBER_FILL if a.get("severity") in ("HIGH","MEDIUM")
                    else LIGHT_FILL)
        row_vals = [
            a.get("detected_at",""), a.get("table",""),
            a.get("severity",""),   a.get("metric",""),
            a.get("detector",""),   a.get("today",""),
            a.get("expected",""),   a.get("score",""),
            a.get("explanation",""),
        ]
        for j, val in enumerate(row_vals, 1):
            c = ws3.cell(row=i, column=j, value=val)
            c.border = thin
            c.alignment = Alignment(vertical="center", wrap_text=(j==9))
            if j == 3:
                c.fill = sev_fill
                c.font = Font(bold=True, size=9)
        ws3.row_dimensions[i].height = 30
    auto_width(ws3)
    ws3.column_dimensions["I"].width = 60

    # ── Sheet 4: Alert Log ────────────────────────────────────
    if alert_log:
        ws4 = wb.create_sheet("Alert Log")
        headers4 = ["Sent At","Channel","Severity","Summary"]
        header_row(ws4, 1, headers4)
        for i, a in enumerate(alert_log, 2):
            row_vals = [a.get("sent_at",""), a.get("channel",""),
                        a.get("severity",""), a.get("summary","")]
            for j, val in enumerate(row_vals, 1):
                c = ws4.cell(row=i, column=j, value=val)
                c.border = thin
                if i % 2 == 1:
                    c.fill = LIGHT_FILL
        auto_width(ws4)

    # ── Sheet 5: Row Counts (per table) ───────────────────────
    if charts:
        ws5 = wb.create_sheet("Row Count History")
        col = 1
        for table, data in charts.items():
            dates  = data.get("dates", [])
            counts = data.get("row_counts", [])
            ws5.cell(1, col, table).font = font(bold=True)
            ws5.cell(2, col, "Date")
            ws5.cell(2, col+1, "Row Count")
            for j, (d, c) in enumerate(zip(dates, counts), 3):
                ws5.cell(j, col, d)
                ws5.cell(j, col+1, c)
            col += 3
        auto_width(ws5)

    wb.save(output_path)
    print(f"  ✓ Excel: {output_path}")


# ════════════════════════════════════════════════════════════════
#  MAIN
# ════════════════════════════════════════════════════════════════
def main():
    pdf_only   = "--pdf-only"   in sys.argv
    csv_only   = "--csv-only"   in sys.argv
    excel_only = "--excel-only" in sys.argv
    all_formats = not (pdf_only or csv_only or excel_only)

    output_dir = "./reports"
    if "--output" in sys.argv:
        idx = sys.argv.index("--output")
        if idx + 1 < len(sys.argv):
            output_dir = sys.argv[idx + 1]

    os.makedirs(output_dir, exist_ok=True)
    date_str = datetime.now().strftime("%Y-%m-%d")
    base     = f"{output_dir}/dq_report_{date_str}"

    print(f"\nDQ Monitor — Report Generator")
    print(f"Output directory: {output_dir}/\n")

    summary, dq_scores, anomalies, alert_log, charts, sources = load_data()

    if not summary:
        print("ERROR: No data found. Run python export_dashboard_data.py first.")
        sys.exit(1)

    status = summary.get("overall_status","UNKNOWN")
    print(f"  Status:     {status}")
    print(f"  Tables:     {len(dq_scores)}")
    print(f"  Anomalies:  {len(anomalies)}")
    print(f"  Alerts:     {len(alert_log)}")
    print()

    if all_formats or pdf_only:
        generate_pdf(f"{base}.pdf", summary, dq_scores,
                     anomalies, alert_log, charts, sources)

    if all_formats or csv_only:
        generate_csv(f"{base}.csv", anomalies, dq_scores)

    if all_formats or excel_only:
        try:
            import openpyxl
            generate_excel(f"{base}.xlsx", summary, dq_scores,
                           anomalies, alert_log, charts)
        except ImportError:
            print("  ⚠ Excel skipped — run: pip install openpyxl")

    print(f"\nAll reports saved to: {output_dir}/")
    print(f"Open PDF: {base}.pdf\n")


if __name__ == "__main__":
    main()
